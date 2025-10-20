import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
from sklearn.linear_model import LinearRegression

# ==============================
# 1. Load and clean dataset
# ==============================
rdf = pd.read_csv(r"C:\Users\osama\Desktop\airport.as\los-angeles-international-airport-air-cargo-volume.csv")

print("Initial shape:", rdf.shape)
print("Columns:", rdf.columns.tolist())
print("Missing values:\n", rdf.isna().sum())

rdf = rdf.drop_duplicates().dropna()
rdf["ReportPeriod"] = pd.to_datetime(rdf["ReportPeriod"], errors="coerce")

rdf["Year"] = rdf["ReportPeriod"].dt.year
rdf["Month"] = rdf["ReportPeriod"].dt.month
rdf["CargoTons_K"] = rdf["AirCargoTons"] / 1000

print(f"\nData covers years {rdf['Year'].min()} to {rdf['Year'].max()}")




import pandas as pd

csv_path = r"C:\Users\osama\Desktop\airport.as\los-angeles-international-airport-air-cargo-volume.csv"
rdf = pd.read_csv(csv_path)

print("✅ Data loaded successfully! Rows:", len(rdf))
print(rdf.head())





# Create output directory
output_dir = r"C:\Users\osama\Desktop\airport.as\plots"
os.makedirs(output_dir, exist_ok=True)

# ==============================
# 2. Annual Analysis
# ==============================
import pandas as pd

csv_path = r"C:\Users\osama\Desktop\airport.as\los-angeles-international-airport-air-cargo-volume.csv"
rdf = pd.read_csv(csv_path)

print("✅ Data loaded successfully! Rows:", len(rdf))
print(rdf.head())

# 🧩 تحويل الأعمدة الزمنية
rdf["ReportPeriod"] = pd.to_datetime(rdf["ReportPeriod"], errors="coerce")

# إنشاء عمود السنة والشهر
rdf["Year"] = rdf["ReportPeriod"].dt.year
rdf["Month"] = rdf["ReportPeriod"].dt.month

# تحويل الكمية من طن إلى آلاف الأطنان (اختياري، لأجل الرسم)
rdf["CargoTons_K"] = rdf["AirCargoTons"] / 1000

# 👇 الآن نقدر نجمع حسب السنة
annual_cargo = rdf.groupby("Year")["CargoTons_K"].sum().reset_index()

print("\n📊 Annual cargo summary:")
print(annual_cargo.head())


plt.figure(figsize=(10, 6))
plt.bar(annual_cargo["Year"], annual_cargo["CargoTons_K"], color="skyblue")
plt.title("Annual Air Cargo Volume - Los Angeles International Airport")
plt.xlabel("Year")
plt.ylabel("Cargo Volume (thousands of tons)")
plt.grid(axis='y')
plt.tight_layout()
annual_path = os.path.join(output_dir, "annual_air_cargo_volume.png")
plt.savefig(annual_path)
plt.show()
print("Saved:", annual_path)

# ==============================
# 3. Monthly Analysis
# ==============================
month_by_cargo = rdf.groupby("Month")["CargoTons_K"].sum().reset_index()

plt.figure(figsize=(10, 6))
plt.bar(month_by_cargo["Month"], month_by_cargo["CargoTons_K"], color="orange")
plt.title("Monthly Air Cargo Volume - Los Angeles International Airport")
plt.xlabel("Month")
plt.ylabel("Cargo Volume (thousands of tons)")
plt.grid(axis='y')
plt.tight_layout()
month_path = os.path.join(output_dir, "monthly_air_cargo_volume.png")
plt.savefig(month_path)
plt.show()
print("Saved:", month_path)

# ==============================
# 4. Domestic vs International
# ==============================
domestic_international = rdf.groupby("Domestic_International")["CargoTons_K"].sum().reset_index()
print("\nDomestic vs International cargo (thousands of tons):")
print(domestic_international)

plt.figure(figsize=(7, 7))
plt.pie(domestic_international["CargoTons_K"], labels=domestic_international["Domestic_International"],
        autopct="%1.1f%%", startangle=140)
plt.title("Domestic vs International Air Cargo Volume")
pie1_path = os.path.join(output_dir, "domestic_vs_international.png")
plt.savefig(pie1_path)
plt.show()
print("Saved:", pie1_path)

# ==============================
# 5. Arrival vs Departure
# ==============================
arrival_departure = rdf.groupby("Arrival_Departure")["CargoTons_K"].sum().reset_index()

plt.figure(figsize=(7, 7))
plt.pie(arrival_departure["CargoTons_K"], labels=arrival_departure["Arrival_Departure"],
        autopct="%1.1f%%", startangle=140)
plt.title("Arrival vs Departure Air Cargo Volume")
pie2_path = os.path.join(output_dir, "arrival_vs_departure.png")
plt.savefig(pie2_path)
plt.show()
print("Saved:", pie2_path)

# ==============================
# 6. Growth Rate
# ==============================
annual_cargo["Growth_Rate"] = annual_cargo["CargoTons_K"].pct_change() * 100

plt.figure(figsize=(10, 6))
plt.plot(annual_cargo["Year"], annual_cargo["Growth_Rate"], marker='o', color='green')
plt.title("Annual Cargo Growth Rate - LAX")
plt.xlabel("Year")
plt.ylabel("Growth Rate (%)")
plt.grid()
growth_path = os.path.join(output_dir, "annual_growth_rate.png")
plt.savefig(growth_path)
plt.show()
print("Saved:", growth_path)

# ==============================
# 7. Forecasting Future Cargo (Simple ML)
# ==============================
X = annual_cargo[["Year"]]
y = annual_cargo["CargoTons_K"]

model = LinearRegression()
model.fit(X, y)

future_years = pd.DataFrame({"Year": np.arange(rdf["Year"].max() + 1, rdf["Year"].max() + 6)})
future_years["Predicted_CargoTons_K"] = model.predict(future_years)

plt.figure(figsize=(10, 6))
plt.plot(annual_cargo["Year"], annual_cargo["CargoTons_K"], marker='o', label="Actual")
plt.plot(future_years["Year"], future_years["Predicted_CargoTons_K"], marker='x', linestyle="--", color="red", label="Forecast")
plt.title("Forecast: Future Air Cargo Volume at LAX")
plt.xlabel("Year")
plt.ylabel("Cargo Volume (thousands of tons)")
plt.legend()
plt.grid()
forecast_path = os.path.join(output_dir, "forecast_cargo_volume.png")
plt.savefig(forecast_path)
plt.show()
print("Saved:", forecast_path)

forecast_df = future_years.copy()
print("\nFuture predictions:\n", forecast_df)




import matplotlib.pyplot as plt
import seaborn as sns

# تأكد إن الأعمدة رقمية
rdf["Year"] = rdf["Year"].astype(int)
rdf["Month"] = rdf["Month"].astype(int)

# خذ الأعمدة المهمة
corr_data = rdf[["Year", "Month", "CargoTons_K"]]

# احسب الارتباط
corr_matrix = corr_data.corr()

print("📈 Correlation matrix:")
print(corr_matrix)

# ارسم heatmap
plt.figure(figsize=(6,4))
sns.heatmap(corr_matrix, annot=True, cmap="coolwarm", fmt=".2f")
plt.title("Correlation between Year, Month, and Cargo Volume")
plt.show()



import pandas as pd
import matplotlib.pyplot as plt

# تأكد أن العمود الزمني مضبوط كـ datetime
rdf["ReportPeriod"] = pd.to_datetime(rdf["ReportPeriod"])

# رتّب البيانات حسب الوقت
rdf = rdf.sort_values("ReportPeriod")

# ارسم الاتجاه العام
plt.figure(figsize=(12,6))
plt.plot(rdf["ReportPeriod"], rdf["CargoTons_K"], label="Monthly Cargo")
plt.title("Cargo Volume Over Time (LAX)")
plt.xlabel("Date")
plt.ylabel("Cargo Tons (K)")
plt.legend()
plt.show()


rdf["RollingMean"] = rdf["CargoTons_K"].rolling(window=12).mean()

plt.figure(figsize=(12,6))
plt.plot(rdf["ReportPeriod"], rdf["CargoTons_K"], alpha=0.5, label="Original Data")
plt.plot(rdf["ReportPeriod"], rdf["RollingMean"], color="red", label="12-Month Rolling Average")
plt.title("Smoothed Trend in Cargo Volume")
plt.xlabel("Date")
plt.ylabel("Cargo Tons (K)")
plt.legend()
plt.show()


#  Forecasting with Prophet
from prophet import Prophet

# تجهيز البيانات للنموذج
forecast_df = annual_cargo.groupby("Year")["CargoTons_K"].sum().reset_index()
forecast_df = forecast_df.rename(columns={"Year": "ds", "CargoTons_K": "y"})

# بناء نموذج Prophet
model = Prophet(yearly_seasonality=True)
model.fit(forecast_df)

# إنشاء تواريخ مستقبلية (مثلاً 5 سنوات قادمة)
future = model.make_future_dataframe(periods=5, freq='Y')

# التنبؤ
forecast = model.predict(future)

# عرض النتائج
forecast[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].tail()

# رسم التنبؤ
model.plot(forecast)
plt.title("📈 Forecast of Air Cargo (Prophet Model)")
plt.xlabel("Year")
plt.ylabel("Predicted Cargo Tons (K)")
plt.tight_layout()
forecast_path = os.path.join(output_dir, "lax_forecast_prophet.png")
plt.savefig(forecast_path)
print("Saved Prophet Forecast Plot:", forecast_path)
plt.show()

# حفظ التنبؤ في ملف Excel
forecast_output_path = os.path.join(output_dir, "LAX_Cargo_Prophet_Forecast.xlsx")
forecast.to_excel(forecast_output_path, index=False)
print("✅ Prophet forecast saved to:", forecast_output_path)


print("📋 Columns in rdf:", rdf.columns.tolist())




# تحويل العمود للتاريخ (لو مو محول مسبقًا)
rdf['ReportPeriod'] = pd.to_datetime(rdf['ReportPeriod'], errors='coerce')

# استخراج اسم الشهر
rdf['Month_Name'] = rdf['ReportPeriod'].dt.strftime('%b')  # Jan, Feb, ...

# تجميع الشحن حسب الشهر
monthly_cargo = rdf.groupby('Month_Name')['CargoTons_K'].sum().reset_index()

print("\nMonthly Cargo Summary:")
print(monthly_cargo)

# رسم البيانات
plt.figure(figsize=(10,6))
plt.bar(monthly_cargo['Month_Name'], monthly_cargo['CargoTons_K'], color='skyblue')
plt.title("Monthly Air Cargo Volume - LAX")
plt.xlabel("Month")
plt.ylabel("Cargo Volume (thousands of tons)")
plt.grid(axis='y')
plt.tight_layout()
plt.show()



import seaborn as sns

# ==============================
# Pivot data for heatmap
# ==============================
monthly_heatmap = rdf.pivot_table(
    index='Year',      # الصفوف = السنوات
    columns='Month',   # الأعمدة = الأشهر
    values='CargoTons_K',  # القيم = حجم الشحن بالآلاف
    aggfunc='sum'      # جمع حجم الشحن لكل شهر
)

print(monthly_heatmap)

# ==============================
# Plot heatmap
# ==============================
plt.figure(figsize=(12, 8))
sns.heatmap(
    monthly_heatmap,
    annot=True,        # كتابة القيم داخل الخلايا
    fmt=".0f",         # بدون كسور عشرية
    cmap="YlGnBu",     # ألوان جذابة من أصفر لأزرق
    linewidths=0.5
)
plt.title("Monthly Air Cargo Volume Heatmap - LAX")
plt.xlabel("Month")
plt.ylabel("Year")
plt.tight_layout()
plt.savefig("monthly_cargo_heatmap.png")
plt.show()



from prophet import Prophet

# ==============================
# Prepare data for forecasting
# ==============================
forecast_df = rdf.groupby("ReportPeriod")["CargoTons_K"].sum().reset_index()
forecast_df = forecast_df.rename(columns={"ReportPeriod": "ds", "CargoTons_K": "y"})

# إنشاء نموذج Prophet
model = Prophet(yearly_seasonality=True, daily_seasonality=False)
model.fit(forecast_df)

# إنشاء بيانات مستقبلية (مثلاً 5 سنوات قادمة)
future = model.make_future_dataframe(periods=60, freq='M')  # 60 شهر = 5 سنوات
forecast = model.predict(future)

# ==============================
# عرض النتائج
# ==============================
plt.figure(figsize=(12, 8))
model.plot(forecast)
plt.title("📈 Forecasted Air Cargo Volume - LAX (Prophet Model)")
plt.xlabel("Date")
plt.ylabel("Cargo Volume (thousands of tons)")
plt.tight_layout()
plt.savefig("lax_forecast_prophet.png")
plt.show()

# ==============================
# عرض التنبؤ في الجدول
# ==============================
forecast_data = forecast[["ds", "yhat", "yhat_lower", "yhat_upper"]]
print("Future forecast (next months):")
print(forecast_data.tail(12))  # آخر 12 شهر من التوقع



# ==============================
# Forecast evaluation
# ==============================
forecast_train = forecast[forecast['ds'] <= forecast_df['ds'].max()]  # توقع على البيانات الحقيقية
merged = forecast_train.merge(forecast_df, left_on='ds', right_on='ds')

# احسب خطأ النموذج
merged['error'] = merged['y'] - merged['yhat']
merged['abs_error'] = merged['error'].abs()
merged['pct_error'] = merged['abs_error'] / merged['y'] * 100

print("Forecast evaluation metrics:")
print("MAE:", merged['abs_error'].mean())
print("MAPE (%):", merged['pct_error'].mean())



# Monthly seasonality
rdf['Month'] = rdf['ReportPeriod'].dt.month
monthly_cargo = rdf.groupby('Month')['CargoTons_K'].sum().reset_index()

plt.figure(figsize=(10, 6))
plt.bar(monthly_cargo['Month'], monthly_cargo['CargoTons_K'], color='orange')
plt.title("📊 Monthly Air Cargo Volume - LAX")
plt.xlabel("Month")
plt.ylabel("Cargo Volume (thousands of tons)")
plt.grid(axis='y')
plt.tight_layout()
plt.show()



# Annual trend
annual_cargo = rdf.groupby('Year')['CargoTons_K'].sum().reset_index()

plt.figure(figsize=(10, 6))
plt.plot(annual_cargo['Year'], annual_cargo['CargoTons_K'], marker='o', color='green')
plt.title("📈 Annual Air Cargo Volume - LAX")
plt.xlabel("Year")
plt.ylabel("Cargo Volume (thousands of tons)")
plt.grid()
plt.tight_layout()
plt.show()



import os

output_dir = r"C:\Users\osama\Desktop\airport.as\plots"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)



excel_path = os.path.join(output_dir, "LAX_Cargo_Analysis.xlsx")

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    
    # 🔹 بيانات سنوية
    annual_cargo.to_excel(writer, sheet_name='Annual_Cargo', index=False)
    
    # 🔹 بيانات شهرية
    monthly_cargo.to_excel(writer, sheet_name='Monthly_Cargo', index=False)
    
    # 🔹 Domestic vs International
    dom_int_col = 'Domestic_International'
    rdf[dom_int_col] = rdf[dom_int_col].astype(str).str.strip()
    dom_int = rdf.groupby(dom_int_col)['CargoTons_K'].sum().reset_index()
    dom_int.to_excel(writer, sheet_name='Domestic_vs_International', index=False)
    
    # 🔹 Arrival vs Departure
    arr_dep_col = 'Arrival_Departure'
    rdf[arr_dep_col] = rdf[arr_dep_col].astype(str).str.strip()
    arr_dep = rdf.groupby(arr_dep_col)['CargoTons_K'].sum().reset_index()
    arr_dep.to_excel(writer, sheet_name='Arrival_vs_Departure', index=False)
    
    # 🔹 نمو الشحن السنوي
    annual_cargo['Growth_Rate'] = annual_cargo['CargoTons_K'].pct_change() * 100
    annual_cargo[['Year', 'CargoTons_K', 'Growth_Rate']].to_excel(writer, sheet_name='Annual_Growth', index=False)
    
    # 🔹 توقع Prophet
    forecast_df.to_excel(writer, sheet_name='Forecast', index=False)
    
print("✅ Excel file created successfully:", excel_path)



import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os

# ==============================
# Load data
# ==============================
rdf = pd.read_csv(r"C:\Users\osama\Desktop\airport.as\los-angeles-international-airport-air-cargo-volume.csv")
rdf = rdf.drop_duplicates().dropna()
rdf["ReportPeriod"] = pd.to_datetime(rdf["ReportPeriod"], errors='coerce')
rdf["Year"] = rdf["ReportPeriod"].dt.year
rdf["Month"] = rdf["ReportPeriod"].dt.month
rdf["CargoTons_K"] = rdf["AirCargoTons"] / 1000

# ==============================
# Create output directory
# ==============================
output_dir = r"C:\Users\osama\Desktop\airport.as\plots"
os.makedirs(output_dir, exist_ok=True)

# ==============================
# Annual cargo
# ==============================
annual_cargo = rdf.groupby("Year")["CargoTons_K"].sum().reset_index()
plt.figure(figsize=(10,6))
plt.bar(annual_cargo["Year"], annual_cargo["CargoTons_K"], color='skyblue')
plt.title("Annual Air Cargo Volume - LAX")
plt.xlabel("Year")
plt.ylabel("Cargo (Thousands of tons)")
plt.grid(True)
annual_path = os.path.join(output_dir, "annual_cargo.png")
plt.savefig(annual_path)
plt.close()

# ==============================
# Monthly cargo
# ==============================
monthly_cargo = rdf.groupby("Month")["CargoTons_K"].sum().reset_index()
plt.figure(figsize=(10,6))
plt.bar(monthly_cargo["Month"], monthly_cargo["CargoTons_K"], color='orange')
plt.title("Monthly Air Cargo Volume - LAX")
plt.xlabel("Month")
plt.ylabel("Cargo (Thousands of tons)")
plt.grid(True)
monthly_path = os.path.join(output_dir, "monthly_cargo.png")
plt.savefig(monthly_path)
plt.close()

# ==============================
# Domestic vs International
# ==============================
dom_int = rdf.groupby("Domestic_International")["CargoTons_K"].sum().reset_index()
plt.figure(figsize=(6,6))
plt.pie(dom_int["CargoTons_K"], labels=dom_int["Domestic_International"], autopct='%1.1f%%', startangle=140)
plt.title("Domestic vs International")
dom_int_path = os.path.join(output_dir, "domestic_international.png")
plt.savefig(dom_int_path)
plt.close()

# ==============================
# Arrival vs Departure
# ==============================
arr_dep = rdf.groupby("Arrival_Departure")["CargoTons_K"].sum().reset_index()
plt.figure(figsize=(6,6))
plt.pie(arr_dep["CargoTons_K"], labels=arr_dep["Arrival_Departure"], autopct='%1.1f%%', startangle=140)
plt.title("Arrival vs Departure")
arr_dep_path = os.path.join(output_dir, "arrival_departure.png")
plt.savefig(arr_dep_path)
plt.close()

# ==============================
# Create Excel Workbook
# ==============================
excel_path = os.path.join(output_dir, "LAX_Cargo_Dashboard.xlsx")
wb = Workbook()
ws = wb.active
ws.title = "Dashboard"

# Insert images
images = [("Annual Cargo", annual_path), 
          ("Monthly Cargo", monthly_path),
          ("Domestic vs International", dom_int_path),
          ("Arrival vs Departure", arr_dep_path)]

row = 1
for title, img_path in images:
    ws.cell(row=row, column=1, value=title)
    img = Image(img_path)
    img.width = 600
    img.height = 400
    ws.add_image(img, f"A{row+1}")
    row += 22  # Adjust for image height

# Save Excel
wb.save(excel_path)
print("✅ Dashboard Excel created:", excel_path)



import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os
from prophet import Prophet

# ==============================
# Load & Clean Data
# ==============================
rdf = pd.read_csv(r"C:\Users\osama\Desktop\airport.as\los-angeles-international-airport-air-cargo-volume.csv")
rdf = rdf.drop_duplicates().dropna()
rdf["ReportPeriod"] = pd.to_datetime(rdf["ReportPeriod"], errors='coerce')
rdf["Year"] = rdf["ReportPeriod"].dt.year
rdf["Month"] = rdf["ReportPeriod"].dt.month
rdf["CargoTons_K"] = rdf["AirCargoTons"] / 1000

# ==============================
# Output Directory
# ==============================
output_dir = r"C:\Users\osama\Desktop\airport.as\plots"
os.makedirs(output_dir, exist_ok=True)

# ==============================
# Annual Cargo
# ==============================
annual_cargo = rdf.groupby("Year")["CargoTons_K"].sum().reset_index()
plt.figure(figsize=(10,6))
plt.bar(annual_cargo["Year"], annual_cargo["CargoTons_K"], color='skyblue')
plt.title("Annual Air Cargo Volume - LAX")
plt.xlabel("Year")
plt.ylabel("Cargo (Thousands of tons)")
plt.grid(True)
annual_path = os.path.join(output_dir, "annual_cargo.png")
plt.savefig(annual_path)
plt.close()

# ==============================
# Monthly Cargo
# ==============================
monthly_cargo = rdf.groupby("Month")["CargoTons_K"].sum().reset_index()
plt.figure(figsize=(10,6))
plt.bar(monthly_cargo["Month"], monthly_cargo["CargoTons_K"], color='orange')
plt.title("Monthly Air Cargo Volume - LAX")
plt.xlabel("Month")
plt.ylabel("Cargo (Thousands of tons)")
plt.grid(True)
monthly_path = os.path.join(output_dir, "monthly_cargo.png")
plt.savefig(monthly_path)
plt.close()

# ==============================
# Domestic vs International
# ==============================
dom_int = rdf.groupby("Domestic_International")["CargoTons_K"].sum().reset_index()
plt.figure(figsize=(6,6))
plt.pie(dom_int["CargoTons_K"], labels=dom_int["Domestic_International"], autopct='%1.1f%%', startangle=140)
plt.title("Domestic vs International")
dom_int_path = os.path.join(output_dir, "domestic_international.png")
plt.savefig(dom_int_path)
plt.close()

# ==============================
# Arrival vs Departure
# ==============================
arr_dep = rdf.groupby("Arrival_Departure")["CargoTons_K"].sum().reset_index()
plt.figure(figsize=(6,6))
plt.pie(arr_dep["CargoTons_K"], labels=arr_dep["Arrival_Departure"], autopct='%1.1f%%', startangle=140)
plt.title("Arrival vs Departure")
arr_dep_path = os.path.join(output_dir, "arrival_departure.png")
plt.savefig(arr_dep_path)
plt.close()

# ==============================
# Forecast using Prophet
# ==============================
prophet_df = rdf.groupby("ReportPeriod")["CargoTons_K"].sum().reset_index()
prophet_df.rename(columns={"ReportPeriod": "ds", "CargoTons_K": "y"}, inplace=True)

model = Prophet(yearly_seasonality=True, weekly_seasonality=False, daily_seasonality=False)
model.fit(prophet_df)

future = model.make_future_dataframe(periods=12, freq='M')
forecast = model.predict(future)

plt.figure(figsize=(10,6))
model.plot(forecast)
plt.title("Prophet Forecast - LAX Air Cargo")
forecast_path = os.path.join(output_dir, "forecast_prophet.png")
plt.savefig(forecast_path)
plt.close()

# ==============================
# Create Excel Dashboard
# ==============================
excel_path = os.path.join(output_dir, "LAX_Cargo_Dashboard.xlsx")
wb = Workbook()
ws = wb.active
ws.title = "Dashboard"

images = [
    ("Annual Cargo", annual_path),
    ("Monthly Cargo", monthly_path),
    ("Domestic vs International", dom_int_path),
    ("Arrival vs Departure", arr_dep_path),
    ("Forecast (Prophet)", forecast_path)
]

row = 1
for title, img_path in images:
    ws.cell(row=row, column=1, value=title)
    img = Image(img_path)
    img.width = 600
    img.height = 400
    ws.add_image(img, f"A{row+1}")
    row += 22

wb.save(excel_path)
print("✅ Dashboard Excel created:", excel_path)




import os
from fpdf import FPDF

# ==============================
# إعداد المسارات
# ==============================
output_dir = r"C:\Users\osama\Desktop\airport.as\plots"
pdf_path = os.path.join(output_dir, "LAX_Cargo_Report.pdf")

# الصور التي أعددناها سابقًا
images = [
    ("Annual Air Cargo Volume", os.path.join(output_dir, "annual_cargo.png")),
    ("Monthly Air Cargo Volume", os.path.join(output_dir, "monthly_cargo.png")),
    ("Domestic vs International", os.path.join(output_dir, "domestic_international.png")),
    ("Arrival vs Departure", os.path.join(output_dir, "arrival_departure.png")),
    ("Prophet Forecast", os.path.join(output_dir, "forecast_prophet.png"))
]

# ==============================
# إعداد PDF
# ==============================
pdf = FPDF(orientation='P', unit='mm', format='A4')
pdf.set_auto_page_break(auto=True, margin=15)
pdf.add_page()
pdf.set_font("Arial", 'B', 16)
pdf.multi_cell(0, 10, "LAX Air Cargo Analysis Report\n\n", align='C')

pdf.set_font("Arial", '', 12)
intro = (
    "This report provides a comprehensive analysis of air cargo volumes at Los Angeles International Airport (LAX).\n"
    "It includes data cleaning, annual and monthly analysis, domestic vs international cargo, arrivals vs departures, "
    "and forecasts using Prophet.\n"
)
pdf.multi_cell(0, 8, intro)

# ==============================
# إضافة الصور والشرح
# ==============================
for title, img_path in images:
    pdf.add_page()
    pdf.set_font("Arial", 'B', 14)
    pdf.multi_cell(0, 8, title + "\n", align='C')
    
    pdf.set_font("Arial", '', 12)
    description = f"This chart illustrates the {title.lower()} at LAX airport.\n"
    pdf.multi_cell(0, 6, description)
    
    if os.path.exists(img_path):
        pdf.image(img_path, x=15, y=None, w=180)

# ==============================
# حفظ PDF
# ==============================
pdf.output(pdf_path)
print("✅ PDF report created successfully:", pdf_path)





import os
import pandas as pd
from fpdf import FPDF

# ==============================
# Paths
# ==============================
output_dir = r"C:\Users\osama\Desktop\airport.as\plots"
pdf_path = os.path.join(output_dir, "LAX_Cargo_Full_Report.pdf")

# CSVs / cleaned data
annual_csv = os.path.join(output_dir, "lax_annual_summary.xlsx")
cleaned_csv = os.path.join(output_dir, "lax_cleaned.csv")

# Charts
charts = [
    ("Annual Air Cargo Volume", os.path.join(output_dir, "annual_air_cargo_volume.png")),
    ("Monthly Air Cargo Volume", os.path.join(output_dir, "monthly_air_cargo_volume.png")),
    ("Domestic vs International", os.path.join(output_dir, "domestic_international_air_cargo_volume.png")),
    ("Arrival vs Departure", os.path.join(output_dir, "arrival_departure_air_cargo_volume.png")),
    ("Annual Cargo Growth Rate", os.path.join(output_dir, "annual_air_cargo_growth_rate.png")),
    ("Prophet Forecast", os.path.join(output_dir, "lax_forecast_prophet.png"))
]

# ==============================
# Load data
# ==============================
annual_cargo = pd.read_excel(annual_csv)
cleaned_df = pd.read_csv(cleaned_csv)

# ==============================
# Setup PDF
# ==============================
pdf = FPDF(orientation='P', unit='mm', format='A4')
pdf.set_auto_page_break(auto=True, margin=15)
pdf.add_page()
pdf.set_font("Arial", 'B', 16)
pdf.multi_cell(0, 10, "LAX Air Cargo Analysis Report\n\n", align='C')

pdf.set_font("Arial", '', 12)
intro = (
    "This report provides a comprehensive analysis of air cargo volumes at Los Angeles International Airport (LAX).\n"
    "It includes data cleaning, annual and monthly analysis, domestic vs international cargo, arrivals vs departures, "
    "growth rates, and forecast using Prophet.\n"
)
pdf.multi_cell(0, 8, intro)

# ==============================
# Add Annual Table
# ==============================
pdf.add_page()
pdf.set_font("Arial", 'B', 14)
pdf.multi_cell(0, 8, "Annual Air Cargo Summary\n", align='C')
pdf.set_font("Arial", '', 11)
for idx, row in annual_cargo.iterrows():
    pdf.multi_cell(0, 6, f"Year: {row['Year']}, Cargo (thousands tons): {row['CargoTons_K']:.2f}, Growth Rate: {row.get('Growth_Rate', 0):.2f}%")
    
# ==============================
# Add Charts
# ==============================
for title, chart_path in charts:
    pdf.add_page()
    pdf.set_font("Arial", 'B', 14)
    pdf.multi_cell(0, 8, title + "\n", align='C')
    
    pdf.set_font("Arial", '', 12)
    pdf.multi_cell(0, 6, f"This chart illustrates the {title.lower()} at LAX airport.\n")
    
    if os.path.exists(chart_path):
        pdf.image(chart_path, x=15, y=None, w=180)

# ==============================
# Save PDF
# ==============================
pdf.output(pdf_path)
print("✅ Professional PDF report created successfully:", pdf_path)












# ==============================
# 8. Save Cleaned + Analysis Results
# ==============================
cleaned_csv_path = os.path.join(output_dir, "lax_cleaned.csv")
rdf.to_csv(cleaned_csv_path, index=False)
print("Cleaned CSV saved to:", cleaned_csv_path)

excel_path = os.path.join(output_dir, "LAX_Cargo_Analysis.xlsx")

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    annual_cargo.to_excel(writer, sheet_name="Annual_Cargo", index=False)
    month_by_cargo.to_excel(writer, sheet_name="Monthly_Cargo", index=False)
    domestic_international.to_excel(writer, sheet_name="Domestic_vs_International", index=False)
    arrival_departure.to_excel(writer, sheet_name="Arrival_vs_Departure", index=False)
    forecast_df.to_excel(writer, sheet_name="Cargo_Forecast", index=False)

print(f"✅ Excel file created successfully: {excel_path}")
